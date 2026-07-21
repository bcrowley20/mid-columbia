"""HOBOconnect app XLSX export handler - used by In-Stream wells.

See Implementation Plan.md section 2 and 6 for the format notes this implements:
readings live on a "Data" sheet, deployment/marker events live on a separate
"Events" sheet (same marker-column convention as the CSV format, just its own
sheet with its own row numbering) - not inline with the data rows as originally
assumed. Timestamps are Excel serial dates representing true DST-aware local
time, requiring zoneinfo conversion rather than a fixed per-file offset. The
vendor's ATM/depth_m/depth_ft columns on the Data sheet are read but discarded
(see section 2/9/10) - depth is always computed ourselves.
"""

from __future__ import annotations

from datetime import datetime, timedelta
from datetime import timezone as dt_timezone
from pathlib import Path
from zoneinfo import ZoneInfo, ZoneInfoNotFoundError

import openpyxl

from ..models import DeploymentEvent, ParameterType, Reading, WellType
from ._util import extract_unit
from .base import LoggerHandler, ParseError

_EXCEL_EPOCH = datetime(1899, 12, 30)

# Events sheet marker column header prefix -> normalized DeploymentEvent.kind
_EVENT_KINDS = {
    "Host Connected": "logger_retrieved",
    "End of File": "end_of_file",
    "Started": "logger_launched",
    "Button Up": "button_up",
    "Button Down": "button_down",
}


def _excel_serial_to_utc(serial: float | datetime, tz: ZoneInfo) -> datetime:
    # openpyxl auto-converts date-formatted numeric cells to naive datetimes;
    # fall back to manual Excel-serial decoding for a plain float, just in case.
    naive_local = serial if isinstance(serial, datetime) else _EXCEL_EPOCH + timedelta(days=serial)
    # fold=0 selects the earlier of the two moments for an ambiguous DST
    # fall-back local time (see Implementation Plan.md section 15).
    localized = naive_local.replace(tzinfo=tz, fold=0)
    return localized.astimezone(dt_timezone.utc)


class _DataColumns:
    def __init__(self, header: tuple):
        self.date_time_index: int | None = None
        self.pressure_index: int | None = None
        self.pressure_unit: str = ""
        self.temp_index: int | None = None
        self.temp_unit: str = ""

        for i, cell in enumerate(header):
            if not isinstance(cell, str):
                continue
            if cell.startswith("Date-Time"):
                self.date_time_index = i
            elif cell.startswith("Absolute Pressure"):
                self.pressure_index = i
                self.pressure_unit = extract_unit(cell)
            elif cell.startswith("Temperature"):
                self.temp_index = i
                self.temp_unit = extract_unit(cell)

        if self.date_time_index is None:
            raise ParseError(f"Data sheet is missing a 'Date-Time' column: {header}")


class _EventColumns:
    def __init__(self, header: tuple):
        self.date_time_index: int | None = None
        self.marker_indices: dict[int, str] = {}

        for i, cell in enumerate(header):
            if not isinstance(cell, str):
                continue
            if cell.startswith("Date-Time"):
                self.date_time_index = i
                continue
            for prefix, kind in _EVENT_KINDS.items():
                if cell.startswith(prefix):
                    self.marker_indices[i] = kind
                    break

        if self.date_time_index is None:
            raise ParseError(f"Events sheet is missing a 'Date-Time' column: {header}")


class HoboConnectXlsxHandler(LoggerHandler):
    name = "hoboconnect_xlsx"

    def can_handle(self, path: Path) -> bool:
        return path.suffix.lower() == ".xlsx"

    def parse(
        self, path: Path, well_id: str, well_type: WellType, timezone: str
    ) -> tuple[list[Reading], list[DeploymentEvent]]:
        try:
            tz = ZoneInfo(timezone)
        except ZoneInfoNotFoundError as exc:
            raise ParseError(f"unknown timezone {timezone!r}") from exc

        workbook = openpyxl.load_workbook(path, read_only=True, data_only=True)
        try:
            if "Data" not in workbook.sheetnames:
                raise ParseError(f"{path.name} has no 'Data' sheet (found: {workbook.sheetnames})")
            readings = self._parse_data_sheet(workbook["Data"], well_id, well_type, tz, path.name)

            events: list[DeploymentEvent] = []
            if "Events" in workbook.sheetnames:
                events = self._parse_events_sheet(workbook["Events"], well_id, tz, path.name)
        finally:
            workbook.close()

        return readings, events

    def _parse_data_sheet(self, worksheet, well_id, well_type, tz, source_file) -> list[Reading]:
        rows = worksheet.iter_rows(values_only=True)
        header = next(rows, None)
        if header is None:
            raise ParseError(f"{source_file}: Data sheet has no header row")
        columns = _DataColumns(header)

        pressure_param = ParameterType.AIR_PRESSURE if well_type is WellType.ATMOSPHERIC else ParameterType.WATER_PRESSURE
        temp_param = ParameterType.AIR_TEMPERATURE if well_type is WellType.ATMOSPHERIC else ParameterType.WATER_TEMPERATURE

        readings: list[Reading] = []
        for row_num, row in enumerate(rows, start=1):
            serial = row[columns.date_time_index] if columns.date_time_index < len(row) else None
            if serial is None:
                continue  # blank template row past the real data

            timestamp_utc = _excel_serial_to_utc(serial, tz)
            source_row = int(row[0]) if row and row[0] is not None else row_num

            if columns.pressure_index is not None and columns.pressure_index < len(row) and row[columns.pressure_index] is not None:
                readings.append(
                    Reading(
                        well_id=well_id,
                        parameter=pressure_param,
                        timestamp_utc=timestamp_utc,
                        value=float(row[columns.pressure_index]),
                        unit=columns.pressure_unit,
                        source_file=source_file,
                        source_row=source_row,
                    )
                )
            if columns.temp_index is not None and columns.temp_index < len(row) and row[columns.temp_index] is not None:
                readings.append(
                    Reading(
                        well_id=well_id,
                        parameter=temp_param,
                        timestamp_utc=timestamp_utc,
                        value=float(row[columns.temp_index]),
                        unit=columns.temp_unit,
                        source_file=source_file,
                        source_row=source_row,
                    )
                )
        return readings

    def _parse_events_sheet(self, worksheet, well_id, tz, source_file) -> list[DeploymentEvent]:
        rows = worksheet.iter_rows(values_only=True)
        header = next(rows, None)
        if header is None:
            return []
        columns = _EventColumns(header)

        events: list[DeploymentEvent] = []
        for row in rows:
            serial = row[columns.date_time_index] if columns.date_time_index < len(row) else None
            if serial is None:
                continue

            timestamp_utc = _excel_serial_to_utc(serial, tz)
            for marker_index, kind in columns.marker_indices.items():
                if marker_index >= len(row):
                    continue
                value = row[marker_index]
                if isinstance(value, str) and value.strip() == "Logged":
                    events.append(DeploymentEvent(well_id=well_id, timestamp_utc=timestamp_utc, kind=kind, source_file=source_file))
        return events
